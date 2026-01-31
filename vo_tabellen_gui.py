import os
import glob
import re
from datetime import datetime
from copy import copy as copy_style

import openpyxl
from openpyxl.styles import Alignment, PatternFill

import tkinter as tk
from tkinter import ttk, filedialog, messagebox


# ============================================================
# FESTE VERZEICHNISSE
# ============================================================

PROTOKOLL_DIR = ""  # optional; wird in der GUI gewählt (leer = .\Protokolle neben EXE)
LAYOUT_DIR = "Layouts"
INTERNAL_HEADER_TEXT = "NUR FÜR DEN INTERNEN DIENSTGEBRAUCH"

__version__ = "2.0.8"

# ============================================================
# Hilfsfunktionen: Merge-sicher schreiben
# ============================================================

def _merged_top_left(ws, row, col):
    """Gibt (min_row, min_col) des Merge-Bereichs zurück, wenn (row,col) in einer Merge-Range liegt."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return None

def set_value_merge_safe(ws, row, col, value):
    """
    Setzt einen Wert auch dann, wenn die Zielzelle eine MergedCell ist.
    In dem Fall wird in die Top-Left-Zelle der Merge-Range geschrieben.
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        tl = _merged_top_left(ws, row, col)
        if tl is None:
            return False
        row, col = tl
    ws.cell(row=row, column=col).value = value
    return True

RAW_SHEET_NAMES = {
    1: "XML-Tab1-Land",
    2: "XML-Tab2-Land",
    3: "XML-Tab3-Land",
    5: "XML-Tab5-Land",
    8: "XML-Tab8-Land",
}

TEMPLATES = {
    1: {"ext": "Tabelle-1-Layout_g.xlsx", "int": "Tabelle-1-Layout_INTERN.xlsx"},
    2: {"ext": "Tabelle-2-Layout_g.xlsx", "int": "Tabelle-2-Layout_INTERN.xlsx"},
    3: {"ext": "Tabelle-3-Layout_g.xlsx", "int": "Tabelle-3-Layout_INTERN.xlsx"},
    5: {"ext": "Tabelle-5-Layout_g.xlsx", "int": "Tabelle-5-Layout_INTERN.xlsx"},
}

PREFIX_TO_TABLE = {
    "Tabelle-1-Land": 1,
    "Tabelle-2-Land": 2,
    "Tabelle-3-Land": 3,
    "Tabelle-5-Land": 5,
}

# ============================================================
# Logging
# ============================================================

class Logger:
    def __init__(self):
        log_dir = PROTOKOLL_DIR.strip() if isinstance(PROTOKOLL_DIR, str) else ""
        if not log_dir:
            log_dir = os.path.join(os.getcwd(), "Protokolle")
        os.makedirs(log_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        self.path = os.path.join(log_dir, f"vo_tabellen_{ts}.log")

    def log(self, msg: str):
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{stamp}] {msg}"
        print(line)
        with open(self.path, "a", encoding="utf-8") as f:
            f.write(line + "\n")


# ============================================================
# Excel Helper
# ============================================================

def is_numeric_like(v):
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        s = v.strip()
        if s in ("-", "X"):
            return True
        s2 = s.replace(".", "").replace(",", "").replace(" ", "")
        return s2.isdigit()
    return False


def _safe_int(v, default: int) -> int:
    try:
        if v is None:
            return int(default)
        return int(v)
    except Exception:
        return int(default)


GER_MONTHS = (
    "Januar", "Februar", "März", "Maerz", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember"
)
PERIOD_TOKENS = ("Quartal", "Halbjahr", "Jahr", "Jahres", "Q1", "Q2", "Q3", "Q4", "H1", "H2", "JJ")


def find_period_text(ws, search_rows=40, search_cols=12):
    """
    Ermittelt den Berichtszeitraum aus der Eingangsdatei.
    Monat/Quartal/Halbjahr: z.B. 'Dezember 2025', '4. Quartal 2025', '1. Halbjahr 2025'
    Jahr: steht oft nur '2025' -> wird zu 'Jahr 2025'
    """
    hits = []
    max_r = min(search_rows, ws.max_row)
    max_c = min(search_cols, ws.max_column)

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if not s:
                continue

            m_year = re.search(r"(20\d{2})", s)
            if not m_year:
                continue
            year = m_year.group(1)

            if any(m in s for m in GER_MONTHS) or any(tok in s for tok in PERIOD_TOKENS):
                hits.append(s)
                continue

            if re.fullmatch(r"\D*" + re.escape(year) + r"\D*", s):
                hits.append(f"Jahr {year}")

    return hits[-1] if hits else None


def extract_stand_from_raw(ws, max_search_rows=80):
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(max_row, max(max_row - max_search_rows, 1) - 1, -1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "Stand:" in v:
                return v.strip()
    return None


def get_merged_secondary_checker(ws):
    merged = list(ws.merged_cells.ranges)

    def is_secondary(row, col):
        for rg in merged:
            if rg.min_row <= row <= rg.max_row and rg.min_col <= col <= rg.max_col:
                return not (row == rg.min_row and col == rg.min_col)
        return False
    return is_secondary


def get_last_data_col(ws, end_row, max_scan_col=30):
    end_row = max(1, end_row)
    last = 1
    max_c = min(max_scan_col, ws.max_column)
    for r in range(1, end_row + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                last = max(last, c)
    return last


def update_footer_with_stand_and_copyright(ws, stand_text):
    max_row = ws.max_row
    max_col = ws.max_column
    current_year = datetime.now().year

    copyright_row = None
    for r in range(max_row, 0, -1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "(C)opyright" in v:
            new_text = re.sub(r"\(C\)opyright\s+\d{4}", f"(C)opyright {current_year}", v)
            ws.cell(row=r, column=1).value = new_text
            copyright_row = r
            break

    if not copyright_row:
        return

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().startswith("Stand:") and r != copyright_row:
                ws.cell(row=r, column=c).value = ""

    if not stand_text:
        return

    stand_col = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=copyright_row, column=c).value
        if isinstance(v, str) and "Stand:" in v:
            stand_col = c
            break
    if stand_col is None:
        stand_col = get_last_data_col(ws, end_row=copyright_row - 1)

    cop_cell = ws.cell(row=copyright_row, column=1)
    tgt = ws.cell(row=copyright_row, column=stand_col)
    tgt.value = stand_text

    tgt.font = copy_style(cop_cell.font)
    tgt.border = copy_style(cop_cell.border)
    tgt.fill = copy_style(cop_cell.fill)
    tgt.number_format = cop_cell.number_format
    tgt.protection = copy_style(cop_cell.protection)
    tgt.alignment = Alignment(
        horizontal="right",
        vertical=cop_cell.alignment.vertical if cop_cell.alignment else "center"
    )

# ============================================================
# Tabelle 8 (_g) – Batch-Verarbeitung 25..28_Tab8_*
# ============================================================

TAB8_FILE_RE = re.compile(r"^(?P<nr>25|26|27|28)_Tab8_(?P<token>.+)\.xlsx$", re.IGNORECASE)

def resolve_layout_path(candidates):
    """Nimmt die erste existierende Layout-Datei aus candidates (relativ zu LAYOUT_DIR)."""
    for name in candidates:
        p = os.path.join(LAYOUT_DIR, name)
        if os.path.exists(p):
            return p
    return None

def parse_tab8_token(token: str):
    """Erkennt Monats-/Quartals-/Halbjahres-/JJ-Token aus Dateinamens-Token.

    Unterstützt beide Varianten:
      - ohne Jahr: Q4, H2
      - mit Jahr:  2025-Q4, 2025-H2
      - Monat:     2025-11
      - Jahr:      2025-JJ
    """
    token = token.strip()

    # Monat: YYYY-MM
    if re.fullmatch(r"\d{4}-\d{2}", token):
        return ("monat", token)

    # Quartal: Q1..Q4 oder YYYY-Qx
    m = re.fullmatch(r"(?:(\d{4})-)?Q([1-4])", token, flags=re.IGNORECASE)
    if m:
        year = m.group(1)
        q = f"Q{m.group(2)}"
        return ("quartal", f"{year}-{q}" if year else q)

    # Halbjahr: H1/H2 oder YYYY-Hx
    m = re.fullmatch(r"(?:(\d{4})-)?H([12])", token, flags=re.IGNORECASE)
    if m:
        year = m.group(1)
        h = f"H{m.group(2)}"
        return ("halbjahr", f"{year}-{h}" if year else h)

    # Jahr: YYYY-JJ
    if re.fullmatch(r"\d{4}-JJ", token, flags=re.IGNORECASE):
        y = token[:4]
        return ("jj", f"{y}-JJ")

    return (None, token)

def get_file_stand_date(paths):
    """Stand-Datum (dd.mm.yyyy) aus Dateizeitstempeln: nimmt den neuesten mtime-Wert."""
    mt = 0.0
    for p in paths:
        try:
            mt = max(mt, os.path.getmtime(p))
        except Exception:
            pass
    if mt <= 0:
        dt = datetime.now()
    else:
        dt = datetime.fromtimestamp(mt)
    return dt.strftime("%d.%m.%Y")

def tab8_find_title_cell(ws_raw):
    """Sucht eine Titelzelle (typisch A3) und gibt den Text zurück."""
    for r in range(1, 25):
        v = ws_raw.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().startswith("8.") and "Unternehmensinsolvenzen" in v:
            return v
    # Fallback: A3
    v = ws_raw.cell(row=3, column=1).value
    return v if isinstance(v, str) else None

def tab8_detect_data_block(ws_raw):
    """Ermittelt (first_data_row, last_data_row, footer_row).

    Besonderheit: In den Eingangsdateien steht in Zeile 1 / Spalte A häufig nur die
    Blattnummer (25/26/27/28). Diese darf NICHT als Datenbeginn interpretiert werden.

    Heuristik:
      1) Suche die Kopfzeile mit "Schl.-" (meist "Schl.-Nr.") in Spalte A.
      2) Datenbeginn = erste numerische Schl.-Nr. NACH dieser Kopfzeile.
      3) Footer = Trennerzeile (—-_) in Spalte A.
      4) Fallbacks, falls Kopfzeile nicht gefunden wird.
    """

    # Footer-Trenner suchen
    footer = None
    for r in range(ws_raw.max_row, 0, -1):
        v = ws_raw.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() and all(ch in "—-_" for ch in v.strip()):
            footer = r
            break
    if footer is None:
        footer = ws_raw.max_row + 1

    # Kopfzeile "Schl.-Nr." suchen (Spalte A)
    header_row = None
    for r in range(1, min(ws_raw.max_row, 80) + 1):
        v = ws_raw.cell(row=r, column=1).value
        if isinstance(v, str):
            s = v.replace("\n", " ").strip().lower()
            if "schl" in s and "nr" in s:
                header_row = r
                break

    # Datenbeginn: erste numerische Schl.-Nr. nach header_row
    first = None
    start_scan = (header_row + 1) if header_row else 1
    for r in range(start_scan, min(footer, ws_raw.max_row + 1)):
        v = ws_raw.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            # 25/26/27/28-Marker in den ersten Zeilen ignorieren
            if header_row is None and r <= 3 and int(v) in (25, 26, 27, 28):
                continue
            first = r
            break
        if isinstance(v, str) and v.strip().isdigit():
            iv = int(v.strip())
            if header_row is None and r <= 3 and iv in (25, 26, 27, 28):
                continue
            first = r
            break

    if first is None:
        # Sehr defensiver Fallback: nimm header_row+1 oder 1
        first = (header_row + 1) if header_row else 1

    last = max(first, footer - 1)
    return first, last, footer


def tab8_find_footnote_start(ws_out):
    """Findet die erste Fußnoten-Zeile im Layout (typisch beginnt Spalte A mit '-').

    Rückgabe: Zeilennummer oder None.
    """
    for r in range(1, ws_out.max_row + 1):
        v = ws_out.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().startswith("-"):
            return r
    return None


def tab8_scan_stand_cells(ws):
    """Liefert alle Zellenkoordinaten (r,c), deren Inhalt 'Stand:' enthält."""
    coords = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "Stand:" in v:
                # Merge-safe: auf Top-Left normalisieren
                tl = _merged_top_left(ws, r, c)
                if tl is not None:
                    coords.append(tl)
                else:
                    coords.append((r, c))
    # uniq + sort
    coords = sorted(set(coords), key=lambda x: (x[0], x[1]))
    return coords


def tab8_normalize_stand(ws, stand_row: int, stand_col: int, stand_ddmmyyyy: str, ref_cell=None):
    """Sorgt dafür, dass 'Stand:' exakt einmal im Blatt vorkommt.

    - löscht alle anderen 'Stand:'-Vorkommen
    - setzt den Stand in (stand_row, stand_col)
    - übernimmt optional Style von ref_cell
    """
    keep = _merged_top_left(ws, stand_row, stand_col) or (stand_row, stand_col)

    # Alle anderen Stand-Zellen leeren
    for (r, c) in tab8_scan_stand_cells(ws):
        if (r, c) != keep:
            set_value_merge_safe(ws, r, c, "")

    # Zielzelle setzen (merge-safe)
    set_value_merge_safe(ws, stand_row, stand_col, f"Stand: {stand_ddmmyyyy}")
    tl = _merged_top_left(ws, stand_row, stand_col)
    rr, cc = tl if tl is not None else (stand_row, stand_col)
    tgt = ws.cell(row=rr, column=cc)

    if ref_cell is not None:
        try:
            tgt.font = copy_style(ref_cell.font)
            tgt.border = copy_style(ref_cell.border)
            tgt.fill = copy_style(ref_cell.fill)
            tgt.number_format = ref_cell.number_format
            tgt.protection = copy_style(ref_cell.protection)
            tgt.alignment = copy_style(ref_cell.alignment)
        except Exception:
            pass

    return (rr, cc)


def tab8_update_footer(ws_out, stand_ddmmyyyy: str):
    """Setzt Copyright-Jahr auf aktuelles Jahr und Stand: dd.mm.yyyy (wie im Layout).

    Rückgabe:
        (stand_row, stand_col) falls eine Stand-Zelle gefunden wurde, sonst None.
    """
    current_year = datetime.now().year

    # Copyright-Zelle suchen (irgendwo im Blatt) und Jahr ersetzen
    found_copyright = False
    for r in range(ws_out.max_row, 0, -1):
        for c in range(1, ws_out.max_column + 1):
            v = ws_out.cell(row=r, column=c).value
            if isinstance(v, str) and "(C)opyright" in v:
                ws_out.cell(row=r, column=c).value = re.sub(
                    r"\(C\)opyright\s+\d{4}",
                    f"(C)opyright {current_year}",
                    v
                )
                found_copyright = True
                break
        if found_copyright:
            break

    # Stand-Zelle suchen und Datum setzen
    for r in range(ws_out.max_row, 0, -1):
        for c in range(ws_out.max_column, 0, -1):
            v = ws_out.cell(row=r, column=c).value
            if isinstance(v, str) and "Stand:" in v:
                ws_out.cell(row=r, column=c).value = f"Stand: {stand_ddmmyyyy}"
                return (r, c)

    return None

def fill_tab8_sheet(ws_out, raw_path, max_data_col: int, logger: Logger):
    """Füllt ein Layout-Blatt mit den Daten aus raw_path (Spaltenbegrenzung: M=13 oder N=14)."""
    wb_raw = openpyxl.load_workbook(raw_path, data_only=True)
    # Sheetwahl: bevorzugt XML-Tab8-Land, sonst erstes Blatt
    if 8 in RAW_SHEET_NAMES and RAW_SHEET_NAMES[8] in wb_raw.sheetnames:
        ws_raw = wb_raw[RAW_SHEET_NAMES[8]]
    else:
        ws_raw = wb_raw[wb_raw.sheetnames[0]]

    title = tab8_find_title_cell(ws_raw)
    if title:
        set_value_merge_safe(ws_out, 3, 1, title)

    # Tabellenüberschrift (Zeile 3) fett wie im Layout-Blatt 1
    try:
        # Wenn A3 Teil eines Merges ist: alle Zellen im Merge fett setzen
        applied = False
        for r in ws_out.merged_cells.ranges:
            if r.min_row <= 3 <= r.max_row and r.min_col <= 1 <= r.max_col:
                for rr in range(r.min_row, r.max_row + 1):
                    for cc in range(r.min_col, r.max_col + 1):
                        cell = ws_out.cell(row=rr, column=cc)
                        cell.font = copy_style(cell.font)
                        cell.font = cell.font.copy(bold=True)
                applied = True
                break
        if not applied:
            c = ws_out.cell(row=3, column=1)
            c.font = copy_style(c.font)
            c.font = c.font.copy(bold=True)
    except Exception:
        pass

    # A1 muss leer sein
    set_value_merge_safe(ws_out, 1, 1, None)

    # Datenblock erkennen (raw) + (out)
    f_raw, l_raw, footer_raw = tab8_detect_data_block(ws_raw)

    # Ziel-Start: erste numerische Schl.-Nr. im Layout (Spalte A)
    f_out = None
    for r in range(1, ws_out.max_row + 1):
        v = ws_out.cell(row=r, column=1).value
        if isinstance(v, (int, float)) or (isinstance(v, str) and str(v).strip().isdigit()):
            f_out = r
            break
    if f_out is None:
        f_out = f_raw

    # Ziel-Footer/Datenende: Begrenzung des Datenbereichs im Layout.
    # Wichtig: Auf Blatt 1 stehen unterhalb der Daten häufig Fußnoten aus der Layoutdatei.
    # Diese dürfen NICHT durch das Löschen/Überschreiben des Datenbereichs verschwinden.
    #
    # Heuristik (kleinstes zutreffendes Ende):
    #   1) Fußnotenstart: erste Zeile, deren Spalte A mit '-' beginnt
    #   2) Copyright-Zeile
    #   3) Trennzeile (—) in Spalte A
    #
    # data_end_out ist EXKLUSIV (range(f_out, data_end_out))
    data_end_out = None

    # 1) Fußnotenstart
    footnote_start = tab8_find_footnote_start(ws_out)
    if footnote_start:
        data_end_out = footnote_start

    # 2) Copyright-Zeile finden
    copyright_row = None
    for r in range(ws_out.max_row, 0, -1):
        for c in range(1, ws_out.max_column + 1):
            v = ws_out.cell(row=r, column=c).value
            if isinstance(v, str) and "(C)opyright" in v:
                copyright_row = r
                break
        if copyright_row:
            break
    if copyright_row:
        data_end_out = min(data_end_out, copyright_row) if data_end_out else copyright_row

    # 3) Trennzeile (—) in Spalte A
    sep_row = None
    for r in range(ws_out.max_row, 0, -1):
        v = ws_out.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() and all(ch in "—-_" for ch in v.strip()):
            sep_row = r
            break
    if sep_row:
        data_end_out = min(data_end_out, sep_row) if data_end_out else sep_row

    if data_end_out is None:
        data_end_out = ws_out.max_row + 1

    # Sicherheit: Datenbereich muss mindestens eine Zeile haben
    if data_end_out <= f_out:
        data_end_out = ws_out.max_row + 1


    is_sec = get_merged_secondary_checker(ws_out)

    # Vorherige Beispielfüllung löschen im Zielbereich (NUR Datenblock, nicht Fußnoten!)
    for rr in range(f_out, data_end_out):

        for cc in range(1, max_data_col + 1):
            if not is_sec(rr, cc):
                set_value_merge_safe(ws_out, rr, cc, None)

    # Kopieren: Zeile für Zeile
    raw_r = f_raw
    out_r = f_out
    while raw_r <= l_raw and out_r < data_end_out:
        for cc in range(1, max_data_col + 1):
            if is_sec(out_r, cc):
                continue
            set_value_merge_safe(ws_out, out_r, cc, ws_raw.cell(row=raw_r, column=cc).value)
        raw_r += 1
        out_r += 1

def process_tab8_in_dir(input_dir: str, out_dir: str, logger: Logger, status_var: tk.StringVar):
    """Findet 25..28_Tab8_*.xlsx und erzeugt _g-Dateien.

    WICHTIG: Bei euch liegen Tab8/Tab9 oft in einem Unterordner 'Tab-8-9'.
    Daher wird sowohl im input_dir als auch in input_dir/Tab-8-9 gesucht.
    """

    search_dirs = [input_dir]
    sub = os.path.join(input_dir, "Tab-8-9")
    if os.path.isdir(sub):
        search_dirs.append(sub)

    candidates = []
    for d in search_dirs:
        candidates.extend(glob.glob(os.path.join(d, "*_Tab8_*.xlsx")))
    tab8 = {}
    for p in candidates:
        base = os.path.splitext(os.path.basename(p))[0]
        m = TAB8_FILE_RE.match(os.path.basename(p))
        if not m:
            continue
        nr = int(m.group("nr"))
        token = m.group("token")
        kind, norm_token = parse_tab8_token(token)
        if kind is None:
            logger.log(f"[TAB8][SKIP] Unbekannter Zeitraum-Token in {os.path.basename(p)}")
            continue
        key = (kind, norm_token)
        tab8.setdefault(key, {})[nr] = p

    if not tab8:
        return

    # Layouts auflösen (TEMPLATES[8] existiert in der stabilen Basis nicht -> niemals referenzieren)
    layout_g = resolve_layout_path(["Layout_Tab8_g.xlsx", "Tabelle-8-Layout_g.xlsx"])
    layout_jj = resolve_layout_path(["Layout_Tab8_JJ_g.xlsx", "Tabelle-8-Layout_JJ_g.xlsx"])

    if not layout_g:
        raise FileNotFoundError("Layout für Tabelle 8 (_g) fehlt. Erwartet z.B. Layout_Tab8_g.xlsx oder Tabelle-8-Layout_g.xlsx in ./Layouts")
    if not layout_jj:
        logger.log("[TAB8][WARN] JJ-Layout fehlt (Layout_Tab8_JJ_g.xlsx). JJ-Dateien werden übersprungen, bis das Layout vorhanden ist.")

    for (kind, token), parts in sorted(tab8.items()):
        needed = [25,26,27,28]
        missing = [n for n in needed if n not in parts]
        if missing:
            logger.log(f"[TAB8][SKIP] Zeitraum {token}: es fehlen Dateien: {', '.join(map(str, missing))}")
            continue

        status_var.set(f"Tabelle 8 ({token})")
        logger.log(f"[TAB8] Zeitraum {token} ({kind}) – baue _g")

        stand = get_file_stand_date([parts[n] for n in needed])

        if kind == "jj":
            if not layout_jj:
                continue
            layout_path = layout_jj
            max_col = 14  # bis N
            out_name = f"Tabelle-8-Land_{token}_g.xlsx"
        else:
            layout_path = layout_g
            max_col = 13  # bis M
            out_name = f"Tabelle-8-Land_{token}_g.xlsx"

        wb_out = openpyxl.load_workbook(layout_path, rich_text=True)
        # Erwartet 4 Sheets (25..28). Wir mappen nach Index, notfalls nach Namen.
        ws_map = {}
        for ws in wb_out.worksheets:
            # Sheetname beginnt typischerweise mit 25_Tab8...
            m2 = re.match(r"^(25|26|27|28)_Tab8_", ws.title)
            if m2:
                ws_map[int(m2.group(1))] = ws

        # Fallback: erste 4 Sheets in Reihenfolge 25..28
        if len(ws_map) < 4 and len(wb_out.worksheets) >= 4:
            for idx, nr in enumerate([25,26,27,28]):
                ws_map.setdefault(nr, wb_out.worksheets[idx])

        # Absicherung: Layout muss 4 Blätter hergeben
        still_missing = [nr for nr in [25, 26, 27, 28] if nr not in ws_map]
        if still_missing:
            raise ValueError(
                "Layout für Tabelle 8 muss 4 Tabellenblätter enthalten (25..28). Fehlend: "
                + ", ".join(map(str, still_missing))
            )

        # Füllen + umbenennen
        stand_row = None  # Stand-Zeile aus Blatt 1 merken (falls Layout Stand nur dort hat)
        for nr in [25,26,27,28]:
            ws_out = ws_map[nr]
            raw_path = parts[nr]
            base = os.path.splitext(os.path.basename(raw_path))[0]
            try:
                ws_out.title = base  # z.B. 25_Tab8_2025-11
            except Exception:
                pass
            fill_tab8_sheet(ws_out, raw_path, max_col, logger)

                    # Footer/Stand je Blatt aktualisieren (Copyright-Jahr + Stand, falls vorhanden)
        for nr in [25,26,27,28]:
            ws_out = ws_map[nr]
            raw_path = parts[nr]
            base = os.path.splitext(os.path.basename(raw_path))[0]
            try:
                ws_out.title = base  # z.B. 25_Tab8_2025-11
            except Exception:
                pass
            fill_tab8_sheet(ws_out, raw_path, max_col, logger)

            # Copyright-Jahr + Stand in vorhandenen Stand-Zellen aktualisieren
            tab8_update_footer(ws_out, stand)

        # Stand normalisieren: exakt 1x pro Blatt und immer in der letzten Spalte (M bzw. N).
        ref_ws = ws_map[25]
        stand_cells = tab8_scan_stand_cells(ref_ws)
        if stand_cells:
            # Stand-Zeile aus dem Layout ableiten (unterstes Vorkommen)
            src_r, src_c = max(stand_cells, key=lambda x: (x[0], x[1]))
            stand_row = src_r
            src_cell = ref_ws.cell(row=src_r, column=src_c)

            # Blatt 1 zuerst normalisieren (liefert die echte Top-Left Koordinate zurück)
            keep_r, keep_c = tab8_normalize_stand(ref_ws, stand_row, max_col, stand, ref_cell=src_cell)
            ref_cell = ref_ws.cell(row=keep_r, column=keep_c)

            # Blätter 2..4: alle Stand-Dopplungen entfernen, Stand an gleicher Stelle setzen
            for nr in [26,27,28]:
                ws_out = ws_map[nr]
                tab8_normalize_stand(ws_out, stand_row, max_col, stand, ref_cell=ref_cell)
        else:
            logger.log("[TAB8][WARN] Keine 'Stand:'-Zelle im Layout gefunden – Stand wird nicht normalisiert. (Bitte Layout prüfen)")

        out_path = os.path.join(out_dir, out_name)
        wb_out.save(out_path)
        logger.log(f"[TAB8] _g -> {out_path}")



def mark_cells_with_1_or_2(ws, col_index, fill):
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_index)
        v = cell.value
        if isinstance(v, (int, float)) and v in (1, 2):
            cell.fill = fill
        elif isinstance(v, str) and v.strip() in ("1", "2"):
            cell.fill = fill


def format_numeric_cells(ws, skip_cols=None):
    if skip_cols is None:
        skip_cols = set()

    pos = "#\\ ###\\ ###\\ ###\\ ###\\ ##0"
    neg = "-\\ " + pos
    fmt = f"{pos};{neg};0"

    for row in ws.iter_rows():
        for cell in row:
            if cell.column in skip_cols:
                continue
            v = cell.value
            if v is None or v in ("-", "X"):
                continue
            if isinstance(v, (int, float)):
                if isinstance(v, float):
                    cell.value = int(round(v))
                cell.number_format = fmt


def format_percent_column(ws, col_index: int):
    pct_fmt = "0.0;-0.0;0"
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_index)
        v = cell.value
        if v is None or v in ("-", "X"):
            continue
        if isinstance(v, (int, float)):
            cell.number_format = pct_fmt


def detect_data_and_footer_tab1(sheet):
    max_row = sheet.max_row
    first_data = None
    for r in range(1, max_row + 1):
        for c in range(2, sheet.max_column + 1):
            if is_numeric_like(sheet.cell(row=r, column=c).value):
                first_data = r
                break
        if first_data:
            break
    if first_data is None:
        first_data = 1

    footnote_start = max_row + 1
    for r in range(1, max_row + 1):
        v = sheet.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().startswith("-"):
            footnote_start = r
            break
    return first_data, footnote_start


def detect_data_and_footer_tab2_3(sheet):
    max_row = sheet.max_row
    first_data = None
    for r in range(1, max_row + 1):
        for c in range(3, sheet.max_column + 1):
            if is_numeric_like(sheet.cell(row=r, column=c).value):
                first_data = r
                break
        if first_data:
            break
    if first_data is None:
        first_data = 1

    footnote_start = max_row + 1
    for r in range(1, max_row + 1):
        v = sheet.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().startswith("-"):
            footnote_start = r
            break
    return first_data, footnote_start


# ============================================================
# Footer: Row aus INTERN nach _g übernehmen (JJ Sonderfall)
# ============================================================

def _find_copyright_row(ws):
    for r in range(ws.max_row, 0, -1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "(C)opyright" in v:
            return r
    return None


def clear_existing_footer_markers(ws):
    """
    Löscht vorhandene Copyright/Stand-Textfelder im Zielblatt, damit wir sauber neu setzen können.
    (Wir löschen nur Inhalte, keine Zeilen/keine Merges.)
    """
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                s = v.strip()
                if "(C)opyright" in s or s.startswith("Stand:"):
                    # Merge-safe leeren
                    set_value_merge_safe(ws, r, c, "")


def copy_footer_row_from_intern(ws_target, ws_intern, row_shift_up: int = 0, logger=None):
    """
    Kopiert die komplette Copyright/Stand-Zeile aus _INTERN in die Ziel-Mappe.
    Optional: row_shift_up (z.B. 2) => Zielzeile = INTERN-Zeile - 2
    Merge-sicher: wir schreiben nur in die Top-Left-Zellen.
    """
    r_src = _find_copyright_row(ws_intern)
    if not r_src:
        if logger:
            logger.log("[WARN] Keine Copyright-Zeile in _INTERN gefunden – Footer-Kopie übersprungen.")
        return

    r_tgt = max(1, r_src - int(row_shift_up))

    is_sec = get_merged_secondary_checker(ws_target)
    max_c = max(ws_intern.max_column, ws_target.max_column)

    for c in range(1, max_c + 1):
        if is_sec(r_tgt, c):
            continue

        src_cell = ws_intern.cell(row=r_src, column=c)
        val = src_cell.value

        set_value_merge_safe(ws_target, r_tgt, c, val)

        tl = _merged_top_left(ws_target, r_tgt, c)
        if tl is not None:
            rr, cc = tl
        else:
            rr, cc = r_tgt, c

        tgt_cell = ws_target.cell(row=rr, column=cc)
        tgt_cell.font = copy_style(src_cell.font)
        tgt_cell.border = copy_style(src_cell.border)
        tgt_cell.fill = copy_style(src_cell.fill)
        tgt_cell.number_format = src_cell.number_format
        tgt_cell.protection = copy_style(src_cell.protection)
        tgt_cell.alignment = copy_style(src_cell.alignment)


# ============================================================
# Verarbeitung Tabelle 1
# ============================================================

def build_table1_workbook(raw_path, layout_path, internal_layout: bool):
    wb_raw = openpyxl.load_workbook(raw_path, data_only=True)
    ws_raw = wb_raw[RAW_SHEET_NAMES[1]]

    period_text = find_period_text(ws_raw)
    stand_text = extract_stand_from_raw(ws_raw)

    wb_out = openpyxl.load_workbook(layout_path)
    ws_out = wb_out[wb_out.sheetnames[0]]

    if internal_layout:
        set_value_merge_safe(ws_out, 1, 1, INTERNAL_HEADER_TEXT)
        set_value_merge_safe(ws_out, 5, 1, period_text)
    else:
        set_value_merge_safe(ws_out, 3, 1, period_text)

    is_sec = get_merged_secondary_checker(ws_out)
    fdr_raw, ft_raw = detect_data_and_footer_tab1(ws_raw)
    fdr_raw = _safe_int(fdr_raw, 1)
    ft_raw  = _safe_int(ft_raw, ws_raw.max_row + 1)
    fdr_out, ft_out = detect_data_and_footer_tab1(ws_out)
    fdr_out = _safe_int(fdr_out, 1)
    ft_out  = _safe_int(ft_out, ws_out.max_row + 1)

    n_rows = min(max(0, ft_raw - fdr_raw), max(0, ft_out - fdr_out))
    max_col_out = ws_out.max_column

    for off in range(n_rows):
        r_raw = fdr_raw + off
        r_out = fdr_out + off
        for c in range(1, max_col_out + 1):
            if is_sec(r_out, c):
                continue
            set_value_merge_safe(ws_out, r_out, c, ws_raw.cell(row=r_raw, column=c).value)

    update_footer_with_stand_and_copyright(ws_out, stand_text)

    format_percent_column(ws_out, 9)
    format_numeric_cells(ws_out, skip_cols={9})

    return wb_out


def process_table1_file(raw_path, output_dir, logger: Logger):
    base = os.path.splitext(os.path.basename(raw_path))[0]
    is_jj = "-JJ" in base

    layout_g = os.path.join(LAYOUT_DIR, TEMPLATES[1]["ext"])
    layout_i = os.path.join(LAYOUT_DIR, TEMPLATES[1]["int"])

    wb_i = build_table1_workbook(raw_path, layout_i, internal_layout=True)
    out_i = os.path.join(output_dir, base + "_INTERN.xlsx")
    wb_i.save(out_i)
    logger.log(f"[T1] INTERN -> {out_i}")

    if is_jj:
        wb_g = openpyxl.load_workbook(raw_path)  # Formate/Merges behalten
        ws_g = wb_g[RAW_SHEET_NAMES[1]]

        wb_raw = openpyxl.load_workbook(raw_path, data_only=True)
        ws_raw = wb_raw[RAW_SHEET_NAMES[1]]
        period_text = find_period_text(ws_raw)
        stand_text = extract_stand_from_raw(ws_raw)

        if period_text:
            s = str(period_text).strip()
            if re.fullmatch(r"\d{4}", s):
                s = f"Jahr {s}"
            set_value_merge_safe(ws_g, 3, 1, s)

        # NEU: vorhandene Footer-Marker entfernen, dann Footer aus INTERN 2 Zeilen höher einfügen
        clear_existing_footer_markers(ws_g)
        ws_intern = wb_i[wb_i.sheetnames[0]]
        copy_footer_row_from_intern(ws_g, ws_intern, row_shift_up=2, logger=logger)

        # Stand + Copyright-Jahr dynamisch aktualisieren
        update_footer_with_stand_and_copyright(ws_g, stand_text)

        # Markierung: Spalte G (7) wenn Wert 1 oder 2
        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        mark_cells_with_1_or_2(ws_g, 7, fill)

        out_g = os.path.join(output_dir, base + "_g.xlsx")
        wb_g.save(out_g)
        logger.log(f"[T1] _g (JJ: Footer 2 Zeilen höher + Markierung) -> {out_g}")

    else:
        wb_g = build_table1_workbook(raw_path, layout_g, internal_layout=False)
        out_g = os.path.join(output_dir, base + "_g.xlsx")
        wb_g.save(out_g)
        logger.log(f"[T1] _g -> {out_g}")


# ============================================================
# Verarbeitung Tabelle 2 & 3
# ============================================================

def build_table2_3_workbook(table_no, raw_path, layout_path, internal_layout: bool):
    wb_raw = openpyxl.load_workbook(raw_path, data_only=True)
    ws_raw = wb_raw[RAW_SHEET_NAMES[table_no]]

    period_text = find_period_text(ws_raw)
    stand_text = extract_stand_from_raw(ws_raw)

    wb_out = openpyxl.load_workbook(layout_path)
    ws_out = wb_out[wb_out.sheetnames[0]]

    is_year = isinstance(period_text, str) and period_text.strip().startswith("Jahr ")

    if internal_layout:
        set_value_merge_safe(ws_out, 1, 1, INTERNAL_HEADER_TEXT)
        if table_no == 2:
            set_value_merge_safe(ws_out, 6, 1, period_text)
        else:
            set_value_merge_safe(ws_out, 5, 1, period_text)
    else:
        if table_no == 2:
            set_value_merge_safe(ws_out, 6 if is_year else 4, 1, period_text)
        else:
            set_value_merge_safe(ws_out, 5 if is_year else 3, 1, period_text)

    START_COL_COPY = 2
    is_sec = get_merged_secondary_checker(ws_out)

    fdr_raw, ft_raw = detect_data_and_footer_tab2_3(ws_raw)
    fdr_raw = _safe_int(fdr_raw, 1)
    ft_raw  = _safe_int(ft_raw, ws_raw.max_row + 1)
    fdr_out, ft_out = detect_data_and_footer_tab2_3(ws_out)
    fdr_out = _safe_int(fdr_out, 1)
    ft_out  = _safe_int(ft_out, ws_out.max_row + 1)

    n_rows = min(max(0, ft_raw - fdr_raw), max(0, ft_out - fdr_out))

    for off in range(n_rows):
        r_raw = fdr_raw + off
        r_out = fdr_out + off
        for c in range(START_COL_COPY, ws_out.max_column + 1):
            if is_sec(r_out, c):
                continue
            set_value_merge_safe(ws_out, r_out, c, ws_raw.cell(row=r_raw, column=c).value)

    update_footer_with_stand_and_copyright(ws_out, stand_text)

    format_percent_column(ws_out, 7)
    format_numeric_cells(ws_out, skip_cols={7})

    return wb_out


def process_table2_or_3_file(table_no, raw_path, output_dir, logger: Logger):
    base = os.path.splitext(os.path.basename(raw_path))[0]
    is_jj = "-JJ" in base

    layout_g = os.path.join(LAYOUT_DIR, TEMPLATES[table_no]["ext"])
    layout_i = os.path.join(LAYOUT_DIR, TEMPLATES[table_no]["int"])

    wb_i = build_table2_3_workbook(table_no, raw_path, layout_i, internal_layout=True)
    out_i = os.path.join(output_dir, base + "_INTERN.xlsx")
    wb_i.save(out_i)
    logger.log(f"[T{table_no}] INTERN -> {out_i}")

    if is_jj:
        ws = wb_i[wb_i.sheetnames[0]]
        ws.cell(row=1, column=1).value = None

        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        mark_cells_with_1_or_2(ws, 5, fill)

        out_g = os.path.join(output_dir, base + "_g.xlsx")
        wb_i.save(out_g)
        logger.log(f"[T{table_no}] _g (JJ=INTERN ohne Kopf + Markierung) -> {out_g}")
    else:
        wb_g = build_table2_3_workbook(table_no, raw_path, layout_g, internal_layout=False)
        out_g = os.path.join(output_dir, base + "_g.xlsx")
        wb_g.save(out_g)
        logger.log(f"[T{table_no}] _g -> {out_g}")


# ============================================================
# Verarbeitung Tabelle 5
# ============================================================

def build_table5_workbook(raw_path, layout_path, internal_layout: bool, is_jj: bool):
    wb_raw = openpyxl.load_workbook(raw_path, data_only=True)
    ws_raw = wb_raw[RAW_SHEET_NAMES[5]]

    period_text = find_period_text(ws_raw)
    stand_text = extract_stand_from_raw(ws_raw)

    wb_out = openpyxl.load_workbook(layout_path)
    max_row_raw = ws_raw.max_row

    starts = []
    for r in range(1, max_row_raw + 1):
        v = ws_raw.cell(row=r, column=2).value
        if isinstance(v, str) and re.match(r"Bayern\s+\d\)", v.strip()):
            starts.append(r)

    block_ranges = []
    for i, start in enumerate(starts):
        end = (starts[i + 1] - 1) if i < len(starts) - 1 else max_row_raw
        last_nonempty = start
        for rr in range(start, end + 1):
            if any(ws_raw.cell(row=rr, column=c).value not in (None, "") for c in range(1, 25)):
                last_nonempty = rr
        block_ranges.append((start, last_nonempty))

    def fill_sheet_from_block(ws_out, start_row, end_row):
        is_sec = get_merged_secondary_checker(ws_out)

        first_data_out = None
        for r in range(1, ws_out.max_row + 1):
            if is_numeric_like(ws_out.cell(row=r, column=3).value):
                first_data_out = r
                break
        if first_data_out is None:
            return

        if internal_layout:
            cols = range(3, 11)
        else:
            cols = range(3, 11) if is_jj else range(3, 9)

        raw_r = start_row
        out_r = first_data_out
        while raw_r <= end_row and out_r <= ws_out.max_row:
            for c in cols:
                if is_sec(out_r, c):
                    continue
                set_value_merge_safe(ws_out, out_r, c, ws_raw.cell(row=raw_r, column=c).value)
            raw_r += 1
            out_r += 1

        if (not internal_layout) and (not is_jj):
            for rr in range(first_data_out, out_r):
                for cc in (9, 10):
                    if not is_sec(rr, cc):
                        set_value_merge_safe(ws_out, rr, cc, None)

    for i, (start, end) in enumerate(block_ranges):
        if i >= len(wb_out.worksheets):
            break

        ws = wb_out.worksheets[i]

        if internal_layout:
            set_value_merge_safe(ws, 1, 1, INTERNAL_HEADER_TEXT)
            set_value_merge_safe(ws, 5, 1, period_text)
        else:
            set_value_merge_safe(ws, 3, 1, period_text)

        fill_sheet_from_block(ws, start, end)
        update_footer_with_stand_and_copyright(ws, stand_text)

        format_percent_column(ws, 8)
        format_numeric_cells(ws, skip_cols={8})

    return wb_out


def process_table5_file(raw_path, output_dir, logger: Logger):
    base = os.path.splitext(os.path.basename(raw_path))[0]
    is_jj = "-JJ" in base

    layout_g = os.path.join(LAYOUT_DIR, TEMPLATES[5]["ext"])
    layout_i = os.path.join(LAYOUT_DIR, TEMPLATES[5]["int"])

    wb_i = build_table5_workbook(raw_path, layout_i, internal_layout=True, is_jj=is_jj)
    out_i = os.path.join(output_dir, base + "_INTERN.xlsx")
    wb_i.save(out_i)
    logger.log(f"[T5] INTERN -> {out_i}")

    if is_jj:
        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for ws in wb_i.worksheets:
            ws.cell(row=1, column=1).value = None
            mark_cells_with_1_or_2(ws, 6, fill)

        out_g = os.path.join(output_dir, base + "_g.xlsx")
        wb_i.save(out_g)
        logger.log(f"[T5] _g (JJ=INTERN ohne Kopf + Markierung) -> {out_g}")
    else:
        wb_g = build_table5_workbook(raw_path, layout_g, internal_layout=False, is_jj=is_jj)
        out_g = os.path.join(output_dir, base + "_g.xlsx")
        wb_g.save(out_g)
        logger.log(f"[T5] _g -> {out_g}")


# ============================================================
# Dateisuche & Ausgabeordner
# ============================================================

def find_raw_files(input_dir):
    files = []
    for prefix in PREFIX_TO_TABLE.keys():
        files.extend(glob.glob(os.path.join(input_dir, f"{prefix}_*.xlsx")))
    files = sorted(set(files))
    files = [f for f in files if not f.endswith("_g.xlsx") and not f.endswith("_INTERN.xlsx")]
    return files


def ensure_output_run_folder(base_out_dir: str, input_folder: str) -> str:
    vo_base = os.path.join(base_out_dir, "VÖ-Tabellen")
    os.makedirs(vo_base, exist_ok=True)

    run_dir = os.path.join(vo_base, os.path.basename(input_folder.rstrip("\\/")))
    os.makedirs(run_dir, exist_ok=True)
    return run_dir


# ============================================================
# GUI
# ============================================================

def choose_dir(entry: ttk.Entry):
    d = filedialog.askdirectory()
    if d:
        entry.delete(0, tk.END)
        entry.insert(0, d)


def run_for_one_input_dir(input_dir: str, base_out_dir: str, logger: Logger, status_var: tk.StringVar):
    if not os.path.isdir(input_dir):
        logger.log(f"[SKIP] Eingangspfad existiert nicht: {input_dir}")
        return

    out_dir = ensure_output_run_folder(base_out_dir, input_dir)
    logger.log(f"--- Eingang: {input_dir}")
    logger.log(f"--- Ausgabe: {out_dir}")

    files = find_raw_files(input_dir)
    has_tab8 = bool(glob.glob(os.path.join(input_dir, "*_Tab8_*.xlsx")))
    if not files and not has_tab8:
        logger.log(f"[INFO] Keine passenden Rohdateien gefunden in: {input_dir}")
        return

    if files:
        missing_layouts = []
        for t in (1, 2, 3, 5):
            for k in ("ext", "int"):
                p = os.path.join(LAYOUT_DIR, TEMPLATES[t][k])
                if not os.path.exists(p):
                    missing_layouts.append(p)
        if missing_layouts:
            raise FileNotFoundError("Layout-Dateien fehlen:\n" + "\n".join(missing_layouts))

    for f in files:
        fname = os.path.basename(f)
        table_no = None
        for prefix, tno in PREFIX_TO_TABLE.items():
            if fname.startswith(prefix):
                table_no = tno
                break

        if table_no is None:
            logger.log(f"[SKIP] Unbekanntes Muster: {f}")
            continue

        status_var.set(f"Verarbeite {os.path.basename(input_dir)}: {fname}")
        logger.log(f"[START] {fname}")

        if table_no == 1:
            process_table1_file(f, out_dir, logger)
        elif table_no in (2, 3):
            process_table2_or_3_file(table_no, f, out_dir, logger)
        elif table_no == 5:
            process_table5_file(f, out_dir, logger)

        logger.log(f"[OK]    {fname}")

    # Tabelle 8 (_g): 25..28_Tab8_*.xlsx als Batch (4 Blätter in 1 Datei)
    try:
        process_tab8_in_dir(input_dir, out_dir, logger, status_var)
    except Exception as e:
        logger.log(f"[TAB8][FEHLER] {e}")
        raise



def run_processing(monat_dir, quartal_dir, halbjahr_dir, jahr_dir, base_out_dir, logger: Logger, status_var: tk.StringVar):
    if not base_out_dir:
        messagebox.showerror("Fehler", "Ausgabe-Basisordner muss angegeben werden.")
        return

    jobs = []
    if monat_dir.strip(): jobs.append(("Monat", monat_dir.strip()))
    if quartal_dir.strip(): jobs.append(("Quartal", quartal_dir.strip()))
    if halbjahr_dir.strip(): jobs.append(("Halbjahr", halbjahr_dir.strip()))
    if jahr_dir.strip(): jobs.append(("Jahr", jahr_dir.strip()))

    if not jobs:
        messagebox.showwarning("Hinweis", "Kein Eingangsverzeichnis gesetzt (Monat/Quartal/Halbjahr/Jahr).")
        return

    logger.log("=== START GUI-Lauf ===")
    logger.log(f"Ausgabe-Basis: {base_out_dir}")
    logger.log(f"Layouts:       {os.path.abspath(LAYOUT_DIR)}")
    logger.log("Überschreiben: JA (vorhandene Dateien werden ersetzt)")
    logger.log("Jobs: " + ", ".join([f"{lbl}={path}" for lbl, path in jobs]))

    try:
        for lbl, in_dir in jobs:
            status_var.set(f"Starte: {lbl}")
            logger.log(f"== JOB: {lbl} ==")
            run_for_one_input_dir(in_dir, base_out_dir, logger, status_var)

        status_var.set("Fertig.")
        logger.log("=== ENDE GUI-Lauf ===")
        messagebox.showinfo("Fertig", f"Verarbeitung abgeschlossen.\nProtokoll:\n{logger.path}")
    except Exception as e:
        status_var.set("Fehler – siehe Protokoll.")
        logger.log(f"[FEHLER] {repr(e)}")
        messagebox.showerror("Fehler", f"Es ist ein Fehler aufgetreten:\n{e}\n\nProtokoll:\n{logger.path}")


def start_gui():
    root = tk.Tk()
    root.title("VÖ-Tabellen – GUI v2.0.8 (Tabelle 1/2/3/5/8)")

    frm = ttk.Frame(root, padding=12)
    frm.grid(row=0, column=0, sticky="nsew")

    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    frm.columnconfigure(1, weight=1)

    entries = {}

    def add_row(r, label, key):
        ttk.Label(frm, text=label).grid(row=r, column=0, sticky="w", pady=4)
        e = ttk.Entry(frm, width=90)
        e.grid(row=r, column=1, sticky="we", padx=6)
        ttk.Button(frm, text="Auswählen…", command=lambda: choose_dir(e)).grid(row=r, column=2, sticky="e")
        entries[key] = e

    add_row(0, "Monat – Eingangstabellen (optional):", "monat")
    add_row(1, "Quartal – Eingangstabellen (optional):", "quartal")
    add_row(2, "Halbjahr – Eingangstabellen (optional):", "halbjahr")
    add_row(3, "Jahr – Eingangstabellen (optional):", "jahr")
    add_row(4, "Ausgabe-Basisordner (Pflicht):", "outbase")
    add_row(5, "Protokollordner (optional, leer = .\\Protokolle):", "protokoll")

    status_var = tk.StringVar(value="Bereit.")
    ttk.Label(frm, textvariable=status_var).grid(row=6, column=0, columnspan=3, sticky="w", pady=(10, 0))

    logpath_var = tk.StringVar(value="Protokoll wird geschrieben nach: (noch nicht gestartet)")

    def on_run():
        global PROTOKOLL_DIR
        PROTOKOLL_DIR = entries["protokoll"].get().strip()
        logger = Logger()
        logpath_var.set(f"Protokoll wird geschrieben nach: {logger.path}")

        run_processing(
            entries["monat"].get(),
            entries["quartal"].get(),
            entries["halbjahr"].get(),
            entries["jahr"].get(),
            entries["outbase"].get(),
            logger,
            status_var
        )

    btn_row = ttk.Frame(frm)
    btn_row.grid(row=7, column=0, columnspan=3, sticky="e", pady=10)

    ttk.Button(btn_row, text="Start", command=on_run).grid(row=0, column=0, padx=6)
    ttk.Button(btn_row, text="Schließen", command=root.destroy).grid(row=0, column=1, padx=6)

    ttk.Label(
        frm,
        text="Hinweis: Layouts werden aus .\\Layouts geladen. Ausgaben gehen nach <Basis>\\VÖ-Tabellen\\<Eingangsordnername>.",
    ).grid(row=8, column=0, columnspan=3, sticky="w")

    ttk.Label(frm, textvariable=logpath_var).grid(row=9, column=0, columnspan=3, sticky="w")

    root.mainloop()


if __name__ == "__main__":
    start_gui()
